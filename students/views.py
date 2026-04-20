import io
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator

from .models import Student
from schools.models import PartnerSchool

# Grade points mapping
GRADE_POINTS = {'A': 6, 'B': 5, 'C': 4, 'D': 3, 'E': 2, 'F': 1}

# Available A-Level combinations
COMBINATIONS = [
    {'code': 'PCB', 'subjects': ['Physics', 'Chemistry', 'Biology']},
    {'code': 'PCM', 'subjects': ['Physics', 'Chemistry', 'Mathematics']},
    {'code': 'BCM', 'subjects': ['Biology', 'Chemistry', 'Mathematics']},
    {'code': 'HEG', 'subjects': ['History', 'Economics', 'Geography']},
    {'code': 'HEL', 'subjects': ['History', 'Economics', 'Literature']},
    {'code': 'MEG', 'subjects': ['Mathematics', 'Economics', 'Geography']},
]

def calculate_combination_score(grades, combo):
    """Calculate score for a combination based on student grades"""
    total_score = 0
    max_score = 0
    missing = []
    
    for subject in combo['subjects']:
        grade = grades.get(subject)
        if grade and grade in GRADE_POINTS:
            total_score += GRADE_POINTS[grade]
            max_score += 6
        else:
            missing.append(subject)
    
    if max_score == 0:
        return {'score': 0, 'percentage': 0, 'missing': missing}
    
    percentage = (total_score / max_score) * 100
    return {
        'score': total_score,
        'percentage': round(percentage, 1),
        'missing': missing
    }

@csrf_exempt
def student_list(request):
    """List students for authenticated school"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        school = PartnerSchool.objects.get(admin=request.user)
    except PartnerSchool.DoesNotExist:
        return JsonResponse({'error': 'School profile not found'}, status=404)
    
    students = Student.objects.filter(school=school).order_by('-created_at')
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(students, 20)
    
    try:
        students_page = paginator.page(page)
    except:
        students_page = paginator.page(1)
    
    data = []
    for student in students_page:
        # Calculate best combination
        grades = {
            'Mathematics': student.mathematics,
            'Physics': student.physics,
            'Chemistry': student.chemistry,
            'Biology': student.biology,
            'English': student.english,
            'Geography': student.geography,
            'History': student.history,
            'Economics': student.economics,
            'Literature': student.literature
        }
        
        best_combo = None
        best_percentage = 0
        for combo in COMBINATIONS:
            result = calculate_combination_score(grades, combo)
            if result['percentage'] > best_percentage:
                best_percentage = result['percentage']
                best_combo = combo['code']
        
        data.append({
            'id': str(student.id),
            'name': student.name,
            'grades': {
                'mathematics': student.mathematics,
                'physics': student.physics,
                'chemistry': student.chemistry,
                'biology': student.biology,
                'english': student.english,
                'geography': student.geography,
                'history': student.history,
                'economics': student.economics,
                'literature': student.literature
            },
            'career_interest': student.career_interest,
            'best_combination': best_combo,
            'match_percentage': best_percentage,
            'created_at': student.created_at.isoformat() if student.created_at else None
        })
    
    return JsonResponse({
        'count': paginator.count,
        'total_pages': paginator.num_pages,
        'current_page': int(page),
        'results': data
    })

@csrf_exempt
def student_detail(request, pk):
    """Get single student details"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        school = PartnerSchool.objects.get(admin=request.user)
    except PartnerSchool.DoesNotExist:
        return JsonResponse({'error': 'School profile not found'}, status=404)
    
    try:
        student = Student.objects.get(pk=pk, school=school)
        
        # Calculate best combination
        grades = {
            'Mathematics': student.mathematics,
            'Physics': student.physics,
            'Chemistry': student.chemistry,
            'Biology': student.biology,
            'English': student.english,
            'Geography': student.geography,
            'History': student.history,
            'Economics': student.economics,
            'Literature': student.literature
        }
        
        ranked_combos = []
        for combo in COMBINATIONS:
            result = calculate_combination_score(grades, combo)
            ranked_combos.append({
                'code': combo['code'],
                'percentage': result['percentage'],
                'missing_subjects': result['missing']
            })
        
        ranked_combos.sort(key=lambda x: x['percentage'], reverse=True)
        
        return JsonResponse({
            'id': str(student.id),
            'name': student.name,
            'grades': {
                'mathematics': student.mathematics,
                'physics': student.physics,
                'chemistry': student.chemistry,
                'biology': student.biology,
                'english': student.english,
                'geography': student.geography,
                'history': student.history,
                'economics': student.economics,
                'literature': student.literature
            },
            'career_interest': student.career_interest,
            'ranked_combinations': ranked_combos[:5],  # Top 5
            'best_combination': ranked_combos[0]['code'] if ranked_combos else None,
            'match_percentage': ranked_combos[0]['percentage'] if ranked_combos else 0,
            'guidance_notes': student.guidance_notes,
            'created_at': student.created_at.isoformat() if student.created_at else None
        })
        
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

@csrf_exempt
def create_student(request):
    """Create new student record"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        school = PartnerSchool.objects.get(admin=request.user)
        data = json.loads(request.body)
        
        # Validate required fields
        if not data.get('name'):
            return JsonResponse({'error': 'Student name required'}, status=400)
        
        # Create student
        student = Student.objects.create(
            school=school,
            name=data.get('name'),
            mathematics=data.get('mathematics', ''),
            physics=data.get('physics', ''),
            chemistry=data.get('chemistry', ''),
            biology=data.get('biology', ''),
            english=data.get('english', ''),
            geography=data.get('geography', ''),
            history=data.get('history', ''),
            economics=data.get('economics', ''),
            literature=data.get('literature', ''),
            career_interest=data.get('career_interest', '')
        )
        
        # Calculate best combination
        grades = {
            'Mathematics': student.mathematics,
            'Physics': student.physics,
            'Chemistry': student.chemistry,
            'Biology': student.biology,
            'English': student.english,
            'Geography': student.geography,
            'History': student.history,
            'Economics': student.economics,
            'Literature': student.literature
        }
        
        best_percentage = 0
        best_combo = None
        for combo in COMBINATIONS:
            result = calculate_combination_score(grades, combo)
            if result['percentage'] > best_percentage:
                best_percentage = result['percentage']
                best_combo = combo['code']
        
        student.best_combination = best_combo
        student.match_percentage = best_percentage
        student.save()
        
        return JsonResponse({
            'message': 'Student created successfully',
            'student_id': str(student.id),
            'best_combination': best_combo,
            'match_percentage': best_percentage
        }, status=201)
        
    except PartnerSchool.DoesNotExist:
        return JsonResponse({'error': 'School profile not found'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)

@csrf_exempt
def update_student(request, pk):
    """Update student record"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'PATCH':
        return JsonResponse({'error': 'PATCH method required'}, status=405)
    
    try:
        school = PartnerSchool.objects.get(admin=request.user)
        student = Student.objects.get(pk=pk, school=school)
        data = json.loads(request.body)
        
        # Update allowed fields
        allowed_fields = ['name', 'mathematics', 'physics', 'chemistry', 'biology',
                         'english', 'geography', 'history', 'economics', 'literature',
                         'career_interest']
        
        for field in allowed_fields:
            if field in data:
                setattr(student, field, data[field])
        
        # Recalculate best combination
        grades = {
            'Mathematics': student.mathematics,
            'Physics': student.physics,
            'Chemistry': student.chemistry,
            'Biology': student.biology,
            'English': student.english,
            'Geography': student.geography,
            'History': student.history,
            'Economics': student.economics,
            'Literature': student.literature
        }
        
        best_percentage = 0
        best_combo = None
        for combo in COMBINATIONS:
            result = calculate_combination_score(grades, combo)
            if result['percentage'] > best_percentage:
                best_percentage = result['percentage']
                best_combo = combo['code']
        
        student.best_combination = best_combo
        student.match_percentage = best_percentage
        student.save()
        
        return JsonResponse({
            'message': 'Student updated successfully',
            'best_combination': best_combo,
            'match_percentage': best_percentage
        })
        
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except PartnerSchool.DoesNotExist:
        return JsonResponse({'error': 'School profile not found'}, status=404)

@csrf_exempt
def delete_student(request, pk):
    """Delete student record"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'DELETE':
        return JsonResponse({'error': 'DELETE method required'}, status=405)
    
    try:
        school = PartnerSchool.objects.get(admin=request.user)
        student = Student.objects.get(pk=pk, school=school)
        student.delete()
        
        return JsonResponse({'message': 'Student deleted successfully'})
        
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except PartnerSchool.DoesNotExist:
        return JsonResponse({'error': 'School profile not found'}, status=404)

@csrf_exempt
def bulk_add(request):
    """Bulk add students"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        school = PartnerSchool.objects.get(admin=request.user)
        data = json.loads(request.body)
        students_data = data.get('students', [])
        
        created_count = 0
        errors = []
        
        for student_data in students_data:
            try:
                student = Student.objects.create(
                    school=school,
                    name=student_data.get('name'),
                    mathematics=student_data.get('mathematics', ''),
                    physics=student_data.get('physics', ''),
                    chemistry=student_data.get('chemistry', ''),
                    biology=student_data.get('biology', ''),
                    english=student_data.get('english', ''),
                    geography=student_data.get('geography', ''),
                    history=student_data.get('history', ''),
                    economics=student_data.get('economics', ''),
                    literature=student_data.get('literature', ''),
                    career_interest=student_data.get('career_interest', '')
                )
                
                # Calculate best combination
                grades = {
                    'Mathematics': student.mathematics,
                    'Physics': student.physics,
                    'Chemistry': student.chemistry,
                    'Biology': student.biology,
                    'English': student.english,
                    'Geography': student.geography,
                    'History': student.history,
                    'Economics': student.economics,
                    'Literature': student.literature
                }
                
                best_percentage = 0
                best_combo = None
                for combo in COMBINATIONS:
                    result = calculate_combination_score(grades, combo)
                    if result['percentage'] > best_percentage:
                        best_percentage = result['percentage']
                        best_combo = combo['code']
                
                student.best_combination = best_combo
                student.match_percentage = best_percentage
                student.save()
                
                created_count += 1
                
            except Exception as e:
                errors.append({
                    'name': student_data.get('name', 'Unknown'),
                    'error': str(e)
                })
        
        return JsonResponse({
            'created': created_count,
            'errors': errors
        })
        
    except PartnerSchool.DoesNotExist:
        return JsonResponse({'error': 'School profile not found'}, status=404)


# ========== EXPORT FUNCTIONS (work without authentication for testing) ==========

@csrf_exempt
def export_students_excel(request):
    """Export student records to Excel file (works without login for testing)"""
    # If user is not authenticated, fallback to the first school in database
    if request.user.is_authenticated:
        try:
            school = PartnerSchool.objects.get(admin=request.user)
        except PartnerSchool.DoesNotExist:
            return HttpResponse('School profile not found', status=404)
    else:
        # For demo/testing: use the first school
        school = PartnerSchool.objects.first()
        if not school:
            return HttpResponse('No school found', status=404)

    students = Student.objects.filter(school=school).order_by('name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Report"

    headers = ['Name', 'Mathematics', 'Physics', 'Chemistry', 'Biology', 'English',
               'Geography', 'History', 'Economics', 'Literature', 'Career Interest',
               'Best Combination', 'Match Percentage (%)', 'Guidance Notes']

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.name)
        ws.cell(row=row, column=2, value=student.mathematics or '')
        ws.cell(row=row, column=3, value=student.physics or '')
        ws.cell(row=row, column=4, value=student.chemistry or '')
        ws.cell(row=row, column=5, value=student.biology or '')
        ws.cell(row=row, column=6, value=student.english or '')
        ws.cell(row=row, column=7, value=student.geography or '')
        ws.cell(row=row, column=8, value=student.history or '')
        ws.cell(row=row, column=9, value=student.economics or '')
        ws.cell(row=row, column=10, value=student.literature or '')
        ws.cell(row=row, column=11, value=student.career_interest or '')
        ws.cell(row=row, column=12, value=student.best_combination or '')
        ws.cell(row=row, column=13, value=student.match_percentage or 0)
        ws.cell(row=row, column=14, value=student.guidance_notes or '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="students_{school.name}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@csrf_exempt
def export_students_pdf(request):
    """Export student records to PDF file (works without login for testing)"""
    # If user is not authenticated, fallback to the first school in database
    if request.user.is_authenticated:
        try:
            school = PartnerSchool.objects.get(admin=request.user)
        except PartnerSchool.DoesNotExist:
            return HttpResponse('School profile not found', status=404)
    else:
        # For demo/testing: use the first school
        school = PartnerSchool.objects.first()
        if not school:
            return HttpResponse('No school found', status=404)

    students = Student.objects.filter(school=school).order_by('name')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="students_{school.name}_{datetime.now().strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title = Paragraph(f"Student Report - {school.name}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))

    subtitle_style = styles['Normal']
    subtitle = Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.3 * inch))

    headers = ['Name', 'Best Combo', 'Match %', 'Career']
    data = [headers]

    for student in students:
        data.append([
            student.name,
            student.best_combination or '-',
            f"{student.match_percentage}%" if student.match_percentage else '0%',
            student.career_interest or '-'
        ])

    table = Table(data, colWidths=[2.2*inch, 1.2*inch, 0.8*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
    ]))
    elements.append(table)

    doc.build(elements)
    return response